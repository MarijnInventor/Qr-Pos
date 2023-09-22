from getSettings import productsFile,camsource
import tkinter as tk
import customtkinter
from openpyxl import load_workbook
from tkinter import messagebox as msgb
import numpy as np
import pyzbar.pyzbar as pyzbar
import threading
import cv2


searchRange = [i for i in range(2,10000)]


def addProducts():
    currentRow = 2
    found = 0
    window=customtkinter.CTk()
    window.geometry("700x70")  # Size of the window
    window.title("QR Pos - product add tool")  # Adding a title
    
    customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
    customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green



    wb = load_workbook(productsFile)
    sheet = wb.active
    
    
    def addproduct():
        global currentRow
        global found
        currentRow = 2
        found = 0
        searchForEmptyRow()
       
        print("Writing to database...")
        text = inputname.get()
        pos = sheet['A' + str(currentRow)]
        pos.value = text
       
        text = inputprice.get()
        pos = sheet['B' + str(currentRow)]
        text = text.replace(',','.')
        pos.value = float(text)
        
        text = inputnr.get()
        pos = sheet['C' + str(currentRow)]
        pos.value = int(text)
       
        wb.save(productsFile)
        
        inputname.delete(0, 'end')
        inputprice.delete(0, 'end')
        inputnr.delete(0, 'end')
        inputname.focus()
        
        button.configure(text="Added!")
        window.after(1500, lambda: button.configure(text="Add"))
        print("Product added!")
        
    
    def searchForEmptyRow():
        global currentRow
        global found
        global searchRange
        global currentNr
        print("Searching for empty row...")
        while(found == 0):
            if(currentRow in searchRange):
                cellValue = sheet.cell(row=currentRow, column=1).value
                if(cellValue == None):
                    print("Found! Row " + str(currentRow) + " is empty.")
                    found = 1
                else:
                   currentRow += 1
                   
                    
            else:
                showMessage(message="There is no empty row avalible!",timeout="2000")
                print("There is no empty row avalible!")
                found = 1
                
    def deleteproduct(): #not used yet
        global currentRow
        global found
        currentRow = 2
        found = 0
       
        print("Removing product...")
        text = inputnr.get()
        
        pos = sheet['A' + str(currentRow)]
        pos.value = ""
        pos = sheet['B' + str(currentRow)]
        pos.value = ""
        pos = sheet['C' + str(currentRow)]
        pos.value = ""
       
        wb.save(productsFile)
        
        inputname.delete(0, 'end')
        inputprice.delete(0, 'end')
        inputnr.delete(0, 'end')
        inputname.focus()
        
        print("Product deleted!")
        showMessage(message="\nProduct deleted!",timeout="1500")
        
    
    
    def showMessage(message,timeout):
        root = tk.Tk()
        root.withdraw()
        w = tk.Toplevel(root)
        w.overrideredirect(True)
        w.geometry("300x100+{}+{}".format(
            root.winfo_screenwidth() // 2 - 150, 
            root.winfo_screenheight() // 2 - 50))
        label = tk.Label(w, text=message)
        label.pack(padx=50, pady=20)
        w.after(timeout, w.destroy)
        w.mainloop()
        
        
        
    def scan():
        scanned = False
        cap = cv2.VideoCapture(int(camsource))
        font = cv2.FONT_HERSHEY_PLAIN

        while True: # scanned == False:
            _, frame = cap.read()
            decodedObjects = pyzbar.decode(frame)
            
            for obj in decodedObjects:
                qrContent = obj.data
                print(qrContent)
                inputnr.delete(0, 'end')
                inputnr.insert(0, qrContent)
                inputname.focus()
                scanned = True
                
            cv2.imshow("QR-Pos product add tool", frame)
            key = cv2.waitKey(1)


            if cv2.getWindowProperty('QR-Pos product add tool',cv2.WND_PROP_VISIBLE)<1:
                break
        
        cv2.destroyAllWindows()

    def startScan():
        x = threading.Thread(target=scan, args=())
        x.start()

        
    
    startScan()
    
    label=customtkinter.CTkLabel(window,text='Product name')
    label.place(relx=0.12, rely=0.2,anchor=tk.CENTER)
    inputname=customtkinter.CTkEntry(window)
    inputname.place(relx=0.12, rely=0.5,anchor=tk.CENTER)
     
    label=customtkinter.CTkLabel(window,text='Price')
    label.place(relx=0.37, rely=0.2,anchor=tk.CENTER)
    inputprice=customtkinter.CTkEntry(window)
    inputprice.place(relx=0.37, rely=0.5,anchor=tk.CENTER)
     
    label=customtkinter.CTkLabel(window,text='Product number')
    label.place(relx=0.62, rely=0.2,anchor=tk.CENTER)
    inputnr=customtkinter.CTkEntry(window)
    inputnr.place(relx=0.62, rely=0.5,anchor=tk.CENTER)
    
    button=customtkinter.CTkButton(window,text='Add',command=lambda: addproduct())
    button.place(relx=0.85, rely=0.5,anchor=tk.CENTER)
    
#     button=customtkinter.CTkButton(window,text='Delete',command=lambda: deleteproduct())
#     button.place(relx=0.85, rely=0.7,anchor=tk.CENTER)
    
    
    
    window.bind('<Return>',lambda event:addproduct())
     
     
    window.mainloop()
    
    
if __name__ == "__main__":
    addProducts()
