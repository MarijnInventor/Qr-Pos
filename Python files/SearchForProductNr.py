from settings import productsFile
from openpyxl import load_workbook


wb = load_workbook(productsFile)
sheet = wb.active

def searchForProductNr(qrContent):
    currentScan = int(qrContent)
    currentRow = 2
    found = 0
    cellInfo = "A0"
    price = 0.00
    product = "INVALID"
    searchRange = [i for i in range(1001,5000)]

    while(found == 0):
        if(currentScan in searchRange):
            #print('Qr code is: ' + str(currentScan))
            cellValue = sheet.cell(row = currentRow, column = 3).value
            if(cellValue == currentScan):
                cellInfo = sheet['A' + str(currentRow)].value
                if(str(cellInfo) == "None"):
                    #print("No product found")
                    found = 1
                    currentRow = 1
                else:
                    #print(str(currentScan) + " found in row " + str(currentRow))
                    cellInfo = "A" + str(currentRow)
                    product = sheet[cellInfo].value
                    cellInfo = "B" + str(currentRow)
                    price = sheet[cellInfo].value
                    found = 1
                    currentRow = 1
            else:
                cellInfo = sheet['A' + str(currentRow)].value
                if(cellInfo == "None"):
                    #de errormessage staat in main.py
                    found = 1
                    currentRow = 1
                else:
                    currentRow += 1
        else:
            print("invalid QR code")
            found = 1
    return (price, product)
